import pdfplumber
import os 
import pandas as pd
import re
import pymysql
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
os.chdir('D:\\城市饮用水源水质月报')
with pdfplumber.open('城市集中式生活饮用水水源水质月报（2016年12月）.pdf') as pdf:
    total_pages = len(pdf.pages)
    df = pd.DataFrame()
    for page_number in range(total_pages):
        page = pdf.pages[page_number]
        table = page.extract_table()
        df_page = pd.DataFrame(table)
        df=pd.concat([df,df_page],axis=0,ignore_index=True)
df_1=pd.DataFrame()
page0=pdf.pages[0]
text=page0.extract_text()
s= re.findall('\d\d\d\d.年...',text)
strinfo = re.compile('年')
x= strinfo.sub('-',s[0])
strinfo1= re.compile('月')
sx=strinfo1.sub('',x)
strinfo1= re.compile(' ')
ssx=strinfo1.sub('',sx)
if len(ssx)==6:
    p=ssx[5]
    ssx=ssx.split('-',1)[0]+'-0'+p
date={}
for i in df.index:
    date[i]=ssx+'-01'
date1=pd.Series(date)
df_date=date1.to_frame()
df_date.rename(columns={0:"年月日"},inplace=True)
df_1=pd.concat([df_1,df_date],axis=1)
standard = ['达标','超标']
for c in df.columns:
    df_columns=pd.DataFrame(df[c]) 
    if df[c][0]=="城市名称":
        df_1=pd.concat([df_1,df_columns],axis=1)
    if df[c][0]=="水源\n类型":
        df_1=pd.concat([df_1,df_columns],axis=1)
    if df[c][0]=='水源名称':
        df_1=pd.concat([df_1,df_columns],axis=1)
    if df[c][0]=='达标\n情况':
        df_1=pd.concat([df_1,df_columns],axis=1)
    if df[c][0]=='水质类别':

        df_1=pd.concat([df_1,df_columns],axis=1)
df_1.columns = ['年月日','城市','检测点','水源类别','达标情况','水质等级']
df_1.drop(index=[0,1],inplace=True)
df_1['水质等级']=df_1['水质等级'].astype(str)
for l in df_1.index:
    if len(df_1['水质等级'][l])!=1:
        df_1['水质等级'][l]=None
df_1['水质等级']=df_1['水质等级']+'类'
#加载现有的excel文件
filepath = r"C:\Users\王明杰\Desktop\2016水质报告.xlsx"
workbook = load_workbook(filepath)
#创建新的sheet表
new_sheel = workbook.create_sheet('2016-12')
for r in dataframe_to_rows(df_1,index=False,header=True):
    new_sheel.append(r)
#保存更改到原excel文件
workbook.save(filepath)#导入一年剩下月份的数据
#格式标准